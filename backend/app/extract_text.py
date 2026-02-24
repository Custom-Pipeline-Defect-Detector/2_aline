from __future__ import annotations
import os
from typing import Tuple
from pypdf import PdfReader
import pdfplumber
from docx import Document as DocxDocument
import openpyxl

def extract_text_from_file(path: str, mime: str | None = None) -> Tuple[str, str]:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".pdf":
        return extract_pdf(path), "pdf"
    if ext == ".docx":
        return extract_docx(path), "docx"
    if ext in (".xlsx", ".xlsm"):
        return extract_xlsx(path), "xlsx"
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read(), "text"
    except Exception:
        return "", "unknown"

def extract_pdf(path: str) -> str:
    try:
        reader = PdfReader(path)
        parts = [(p.extract_text() or "") for p in reader.pages]
        text = "\n".join(parts).strip()
        if text:
            return text
    except Exception:
        pass
    try:
        with pdfplumber.open(path) as pdf:
            parts = [(page.extract_text() or "") for page in pdf.pages]
            return "\n".join(parts).strip()
    except Exception:
        return ""

def extract_docx(path: str) -> str:
    doc = DocxDocument(path)
    return "\n".join([p.text for p in doc.paragraphs]).strip()

def extract_xlsx(path: str) -> str:
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for name in wb.sheetnames[:10]:
        ws = wb[name]
        out.append(f"## SHEET: {name}")
        for r in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200), values_only=True):
            row = ["" if v is None else str(v) for v in r[:30]]
            if any(x.strip() for x in row):
                out.append("\t".join(row))
    return "\n".join(out).strip()
